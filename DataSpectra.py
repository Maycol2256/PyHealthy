import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import os
import subprocess
import time
import pyautogui
from supabase import create_client, Client
from typing import Optional
import customtkinter as ctk
import threading
import math
from datetime import datetime
import shutil
import ctypes
from ctypes import wintypes
import glob
import string
import re
from tkcalendar import DateEntry


# === CONFIGURACIÓN ===
CARPETA_BOTONES = "Buttons"
URL_TECFOOD = "https://food.teknisa.com//df/#/df_entrada#dfe11000_lancamento_entrada"
URL_RETIRADA = "https://food.teknisa.com//est/#/est_relatorios#est31100_posicao_de_estoque"

SUPABASE_URL = "https://ulboklgzjriatmaxzpsi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVsYm9rbGd6anJpYXRtYXh6cHNpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjI3ODQxNDAsImV4cCI6MjA3ODM2MDE0MH0.gY6_K4JQoJxPZmdXMIbFZfiJAOdavbg8jDJW1rOUSPk"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# === VARIABLES GLOBALES ===
archivo_excel = None
df_facturas = pd.DataFrame()
productos_por_factura = {}
codigo_clinica = ""
origen_datos = ""
archivo_retirada = None
df_retirada = pd.DataFrame()
unidad_cargada = ""
carpeta_inventario_actual = ""
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
fecha_personalizada = None

# PALETA DE COLORES
PRIMARY = "#3A7BD5"
PRIMARY_HOVER = "#5EA0FF"
ACCENT = "#FFA726" 
ACCENT_HOVER = "#FFB74D"
CARD_BG = "#1C1E23"
TEXT_MAIN = "#FFFFFF"
TEXT_SECOND = "#B0B0B0"
BG = "#0E0F12"

def silenciar_customtkinter():
    import sys
    class NullWriter:
        def write(self, txt):
            pass
        def flush(self):
            pass
    sys.stderr = NullWriter()
    
    
# SISTEMA DE NOTIFICACIONES OPTIMIZADO - CORREGIDO
_toast_lock = threading.Lock()
_active_toasts = []

def _position_toast_window(win, width=380, height=96, margin_x=24, margin_y=24, offset_index=0):
    try:
        root_x = root.winfo_rootx()
        root_y = root.winfo_rooty()
        root_w = root.winfo_width()
        root_h = root.winfo_height()
    except Exception:
        root_x, root_y, root_w, root_h = 0, 0, win.winfo_screenwidth(), win.winfo_screenheight()

    x = root_x + max(root_w - width - margin_x, 0)
    y = root_y + max(root_h - height - margin_y - offset_index * (height + 12), 0)
    win.geometry(f"{width}x{height}+{x}+{y}")

def _reposition_toasts():
    with _toast_lock:
        for idx, t in enumerate(list(_active_toasts)):
            try:
                if t.winfo_exists():
                    _position_toast_window(t, width=t.winfo_width() or 360, height=t.winfo_height() or 100, offset_index=idx)
            except Exception:
                pass

def mostrar_toast(mensaje, tipo="info", duracion=3200, titulo=None):
    estilos = {
        "info":    {"icon": "ℹ️", "bg": "#0b1220", "dot": "#3A7BD5", "border": "#1E3A8A", "shadow": "#1E40AF"},
        "success": {"icon": "✅", "bg": "#07170b", "dot": "#16A34A", "border": "#166534", "shadow": "#15803D"},
        "warning": {"icon": "⚠️", "bg": "#1f1200", "dot": "#F59E0B", "border": "#92400E", "shadow": "#B45309"},
        "error":   {"icon": "❌", "bg": "#2b0e0e", "dot": "#EF4444", "border": "#991B1B", "shadow": "#DC2626"},
    }
    st = estilos.get(tipo, estilos["info"])

    TOAST_W = 360
    TOAST_H = 100

    with _toast_lock:
        offset_index = len(_active_toasts)

    toast = ctk.CTkToplevel(root)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.configure(fg_color=st["bg"])
    
    # Frame principal
    main_frame = ctk.CTkFrame(toast, fg_color=st["bg"], corner_radius=16)
    main_frame.pack(fill="both", expand=True, padx=1, pady=1)
    
    # Frame de sombra
    shadow_frame = ctk.CTkFrame(main_frame, fg_color=st["shadow"], height=4, corner_radius=8)
    shadow_frame.pack(side="bottom", fill="x", padx=8, pady=(0, 2))
    shadow_frame.pack_propagate(False)
    
    # Frame interno
    wrapper = ctk.CTkFrame(main_frame, fg_color=st["bg"], corner_radius=14, width=TOAST_W-12, height=TOAST_H-12)
    wrapper.pack_propagate(False)
    wrapper.pack(fill="both", expand=False, padx=6, pady=(6, 4))

    # Frame del icono
    icon_frame = ctk.CTkFrame(wrapper, fg_color=st["bg"], width=56, height=56)
    icon_frame.pack_propagate(False)
    icon_frame.pack(side="left", padx=(12, 16), pady=12)

    # Canvas para el círculo
    dot_size = 48
    dot = tk.Canvas(icon_frame, width=56, height=56, highlightthickness=0, bg=st["bg"])
    
    circle_x1 = (56 - dot_size) / 2
    circle_y1 = (56 - dot_size) / 2
    circle_x2 = circle_x1 + dot_size
    circle_y2 = circle_y1 + dot_size
    
    dot.create_oval(circle_x1, circle_y1, circle_x2, circle_y2, fill=st["dot"], outline=st["dot"], width=0)
    
    text_x = 56 / 2
    text_y = 56 / 2
    
    dot.create_text(text_x, text_y, text=st["icon"], font=("Segoe UI Emoji", 16), fill="#FFFFFF")
    dot.pack(fill="both", expand=True)

    # Contenedor de texto
    text_container = ctk.CTkFrame(wrapper, fg_color=st["bg"], corner_radius=0)
    text_container.pack(side="left", fill="both", expand=True, pady=12, padx=(0, 12))

    if titulo:
        lbl_title = ctk.CTkLabel(text_container, text=titulo, font=("Segoe UI", 12, "bold"), text_color="#FFFFFF", anchor="w")
        lbl_title.pack(fill="x", pady=(0, 2))
        lbl_msg = ctk.CTkLabel(text_container, text=mensaje, font=("Segoe UI", 11), text_color="#C7CCD1", wraplength=TOAST_W-140, anchor="w", justify="left")
        lbl_msg.pack(fill="both")
    else:
        lbl_msg = ctk.CTkLabel(text_container, text=mensaje, font=("Segoe UI", 12), text_color="#E6E9EE", wraplength=TOAST_W-140, anchor="w", justify="left")
        lbl_msg.pack(fill="both")

    def close_toast(e=None):
        if toast.winfo_exists():
            with _toast_lock:
                if toast in _active_toasts:
                    _active_toasts.remove(toast)
            try:
                toast.destroy()
            except:
                pass
            _reposition_toasts()
    
    # Bind click para cerrar
    for widget in [wrapper, text_container, lbl_msg, icon_frame, dot, main_frame, shadow_frame]:
        try:
            widget.bind("<Button-1>", close_toast)
        except:
            pass
    if titulo:
        try:
            lbl_title.bind("<Button-1>", close_toast)
        except:
            pass

    toast.update_idletasks()
    _position_toast_window(toast, width=TOAST_W, height=TOAST_H, offset_index=offset_index)
    
    try:
        toast.attributes("-alpha", 0.0)
    except:
        pass

    with _toast_lock:
        _active_toasts.append(toast)

    def _animate_in():
        try:
            for i in range(6):
                alpha = (i + 1) / 6
                try:
                    toast.attributes("-alpha", alpha)
                except:
                    pass
                time.sleep(0.02)
        except:
            pass

    def _animate_out_and_destroy():
        time.sleep(duracion / 1000.0)
        try:
            for i in range(5):
                alpha = max(0.0, 1 - (i + 1) / 5)
                try:
                    toast.attributes("-alpha", alpha)
                except:
                    pass
                time.sleep(0.03)
        except:
            pass
        close_toast()

    threading.Thread(target=_animate_in, daemon=True).start()
    threading.Thread(target=_animate_out_and_destroy, daemon=True).start()

# ANIMACIONES OPTIMIZADAS
def simple_button_hover(button, is_enter=True):
    if is_enter:
        if button == btn_excel:
            button.configure(fg_color=ACCENT_HOVER)
        elif button == btn_supabase:
            button.configure(fg_color=PRIMARY_HOVER)
        elif button == btn_iniciar and button.cget("state") == "normal":
            button.configure(fg_color=PRIMARY_HOVER)
        elif button == btn_volver:
            button.configure(fg_color="#7AB8FF")
        elif button == btn_close or button == btn_exit_menu:
            button.configure(fg_color="#FF6B6B")
        elif button == btn_descargar_informes:
            button.configure(fg_color=ACCENT_HOVER)
        elif button == btn_procesar_informes:
            button.configure(fg_color=PRIMARY_HOVER)
    else:
        if button == btn_excel:
            button.configure(fg_color=ACCENT)
        elif button == btn_supabase:
            button.configure(fg_color=PRIMARY)
        elif button == btn_iniciar and button.cget("state") == "normal":
            button.configure(fg_color=PRIMARY)
        elif button == btn_volver:
            button.configure(fg_color="#5EA0FF")
        elif button == btn_close or button == btn_exit_menu:
            button.configure(fg_color="#FF5252")
        elif button == btn_descargar_informes:
            button.configure(fg_color=ACCENT)
        elif button == btn_procesar_informes:
            button.configure(fg_color=PRIMARY)

def simple_item_hover(item, is_enter=True):
    if is_enter:
        item.configure(fg_color="#1E2025")
    else:
        item.configure(fg_color="#141518")

def quick_pulse_animation(widget, pulse_color, duration=0.3):
    def _pulse():
        try:
            original_color = widget.cget("fg_color")
            widget.configure(fg_color=pulse_color)
            time.sleep(duration)
            widget.configure(fg_color=original_color)
        except Exception:
            pass
    threading.Thread(target=_pulse, daemon=True).start()

def fade_transition(widget, target_alpha, duration=0.2):
    def _fade():
        try:
            steps = 4
            current_alpha = widget.attributes("-alpha") if widget.attributes("-alpha") else 1.0
            
            for i in range(steps):
                progress = (i + 1) / steps
                alpha = current_alpha + (target_alpha - current_alpha) * progress
                widget.attributes("-alpha", alpha)
                time.sleep(duration / steps)
        except Exception:
            widget.attributes("-alpha", target_alpha)
    
    threading.Thread(target=_fade, daemon=True).start()

# Confirmación cerrar app OPTIMIZADA
def confirmar_salida(titulo="Confirmar salida", mensaje="¿Deseas salir de la aplicación?"):
    result = {"value": False}
    width, height = 400, 140

    confirm = ctk.CTkToplevel(root)
    confirm.overrideredirect(True)
    confirm.attributes("-topmost", True)
    confirm.configure(fg_color="#0f1724")
    
    confirm.update_idletasks()
    screen_w = confirm.winfo_screenwidth()
    screen_h = confirm.winfo_screenheight()
    x = (screen_w // 2) - (width // 2)
    y = (screen_h // 2) - (height // 2)
    confirm.geometry(f"{width}x{height}+{x}+{y}")
    
    wrapper = ctk.CTkFrame(confirm, fg_color="#0f1724", corner_radius=14)
    wrapper.pack(fill="both", expand=True, padx=8, pady=8)

    barra = ctk.CTkFrame(wrapper, fg_color="#334155", width=8, corner_radius=14)
    barra.pack(side="left", fill="y", padx=(2, 8), pady=6)

    content = ctk.CTkFrame(wrapper, fg_color="#0f1724", corner_radius=8)
    content.pack(side="left", fill="both", expand=True, padx=(0, 8), pady=8)

    lbl_t = ctk.CTkLabel(content, text=titulo, font=("Segoe UI", 13, "bold"), text_color="#FFFFFF")
    lbl_t.pack(anchor="w", padx=8, pady=(6, 0))

    lbl_m = ctk.CTkLabel(
        content,
        text=mensaje,
        font=("Segoe UI", 11),
        text_color="#9CA3AF",
        wraplength=width - 80,
        anchor="w",
        justify="left"
    )
    lbl_m.pack(anchor="w", padx=8, pady=(4, 8))

    btn_frame = ctk.CTkFrame(content, fg_color="#0f1724")
    btn_frame.pack(anchor="e", padx=8, pady=(0, 8))

    def on_confirm():
        result["value"] = True
        fade_transition(confirm, 0.0, 0.15)
        confirm.after(150, confirm.destroy)

    def on_cancel():
        result["value"] = False
        fade_transition(confirm, 0.0, 0.15)
        confirm.after(150, confirm.destroy)

    btn_cancel = ctk.CTkButton(
        btn_frame,
        text="Cancelar",
        width=110,
        command=on_cancel,
        fg_color="#475569",
        hover_color="#64748B"
    )
    btn_cancel.pack(side="right", padx=(10, 0))

    btn_ok = ctk.CTkButton(
        btn_frame,
        text="Salir",
        width=110,
        command=on_confirm,
        fg_color="#DC2626",
        hover_color="#EF4444"
    )
    btn_ok.pack(side="right")
    
    try:
        confirm.attributes("-alpha", 0.0)
    except Exception:
        pass

    fade_transition(confirm, 1.0, 0.15)
    confirm.wait_window()
    return result["value"]



# ========= FUNCIONES CORREGIDAS PARA INVENTARIO =========

# Lista de códigos de clínica estáticos
CODIGOS_CLINICAS = ["0001", "0011", "0024", "0002", "0031", "0014", "0017", "0018", "0003"]
NOMBRES_CLINICAS = {
    "0001": "HEALTHY MATRIZ",
    "0011": "PLANTA IBAGUE",
    "0024": "SEVIN DRUMMOND",
    "0002": "SEATECH",
    "0031": "CLINICA AZUL MEDPLUS",
    "0014": "BRUNE RETAIL",
    "0017": "CARRITO MEDICADIZ",
    "0018": "CARRITO KERALTY",
    "0003": "ASOCAÑA CALI"
}
def crear_carpeta_inventario(fecha_inventario=None):
    """
    Crea la carpeta FINAL donde se guardarán:
    - Los informes descargados
    - El archivo unificado

    Formato final:
    Stock {fecha_inventario} {fecha_hoy}
    """

    # Ruta del escritorio
    escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
    if not os.path.isdir(escritorio):
        escritorio = os.path.join(os.path.expanduser("~"), "Desktop")

    # Convertimos fecha del inventario en formato ideal
    if fecha_inventario:
        fecha_inv_fmt = fecha_inventario.replace("/", "-")
    else:
        fecha_inv_fmt = datetime.now().strftime("%d-%m-%Y")

    fecha_hoy = datetime.now().strftime("%d-%m-%Y")

    # Nombre final de la carpeta
    nombre_carpeta = f"Stock {fecha_inv_fmt} {fecha_hoy}"

    carpeta_final = os.path.join(escritorio, nombre_carpeta)

    os.makedirs(carpeta_final, exist_ok=True)

    return carpeta_final


def obtener_archivo_mas_reciente(carpeta):
    archivos = []
    for nombre in os.listdir(carpeta):
        if nombre.lower().startswith("est31100") and nombre.lower().endswith((".xlsx", ".xls")):
            ruta = os.path.join(carpeta, nombre)
            archivos.append((ruta, os.path.getmtime(ruta)))
    if not archivos:
        return None
    archivos.sort(key=lambda x: x[1], reverse=True)
    return archivos[0][0]


def obtener_carpeta_descargas():

    # ==============================
    # 1. CACHE (si ya se detectó antes)
    # ==============================
    cache_file = "descargas_cache.json"
    try:
        if os.path.exists(cache_file):
            with open(cache_file, "r", encoding="utf8") as f:
                cache = json.load(f)
                ruta_cache = cache.get("descargas")
                if ruta_cache and os.path.isdir(ruta_cache):
                    print(f"📌 Usando ruta guardada (cache): {ruta_cache}")
                    return ruta_cache
    except Exception:
        pass

    # ==============================================
    # 2. RUTA PERSONALIZADA DETECTADA EN TU EMPRESA
    # ==============================================
    rutas_comunes_empresa = [
        r"O:\perfil",
        r"O:\usuarios",
        r"O:\user",
        r"O:\home",
        r"P:\perfil",
        r"P:\usuarios",
        r"S:\perfil",
        r"S:\usuarios"
    ]

    # Buscar patrones típicos corporativos
    for disco in "OPQRSTUVWXYZ":
        raiz = f"{disco}:\\"
        if not os.path.exists(raiz):
            continue

        for root, dirs, files in os.walk(raiz):
            if re.search(r'(Descargas|Downloads)$', root, re.IGNORECASE):
                print(f"🔍 Carpeta de descargas encontrada en red: {root}")
                _guardar_cache_descargas(root)
                return root

    # =============================================================
    # 3. Detectar carpeta redirigida REAL de Windows (shell folders)
    # =============================================================
    try:
        import winreg

        keys = [
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders",
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
        ]

        for key in keys:
            try:
                reg = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key)
                ruta = winreg.QueryValueEx(reg, "Downloads")[0]

                # Expandir variables de entorno (%USERPROFILE%)
                ruta = os.path.expandvars(ruta)

                if os.path.isdir(ruta):
                    print(f"📌 Detectada carpeta redirigida por Windows: {ruta}")
                    _guardar_cache_descargas(ruta)
                    return ruta

            except Exception:
                continue
    except Exception:
        pass

    # ============================
    # 4. Exploración multi-disco
    # ============================
    nombres = ["Descargas", "Downloads"]

    for disco in "CDEFGHIJKLMNOPQRSTUVWXYZ":
        raiz = f"{disco}:\\"
        if not os.path.exists(raiz):
            continue

        for root, dirs, _ in os.walk(raiz):
            for carpeta in dirs:
                if carpeta in nombres:
                    ruta_final = os.path.join(root, carpeta)
                    print(f"🔍 Encontrada: {ruta_final}")
                    _guardar_cache_descargas(ruta_final)
                    return ruta_final

    # ============================
    # 5. Fallback final
    # ============================
    fallback = os.path.join(os.path.expanduser("~"), "Downloads")
    print(f"⚠️ Usando fallback: {fallback}")
    _guardar_cache_descargas(fallback)
    return fallback

def _guardar_cache_descargas(ruta):
    """ Guarda la ruta válida para acelerar detecciones futuras """
    try:
        with open("descargas_cache.json", "w", encoding="utf8") as f:
            json.dump({"descargas": ruta}, f)
    except Exception:
        pass
def limpiar_carpeta_descargas():
    """Elimina archivos Excel antiguos de inventario de la carpeta de descargas"""
    try:
        carpeta_descargas = obtener_carpeta_descargas()
        archivos_eliminados = 0
        
        for archivo in os.listdir(carpeta_descargas):
            ruta_archivo = os.path.join(carpeta_descargas, archivo)
            # Buscar archivos que contengan EST31100 con cualquier extensión
            if ('EST31100' in archivo or 'Inventario_Clinica_' in archivo):
                try:
                    os.remove(ruta_archivo)
                    archivos_eliminados += 1
                    print(f"🗑️ Eliminado archivo antiguo: {archivo}")
                except Exception as e:
                    print(f"⚠️ No se pudo eliminar {archivo}: {e}")
        
        print(f"🧹 Total de archivos antiguos eliminados: {archivos_eliminados}")
        return archivos_eliminados
        
    except Exception as e:
        print(f"⚠️ Error limpiando carpeta descargas: {e}")
        return 0
def obtener_archivo_mas_reciente(carpeta, patrones=None):
    if patrones is None:
        patrones = ['EST31100*', '*.xlsx', '*.xls']
    archivos = []
    for patron in patrones:
        archivos.extend(glob.glob(os.path.join(carpeta, patron)))
    if not archivos:
        return None
    # ordenar por fecha de modificación (más reciente primero)
    archivos = sorted(archivos, key=os.path.getmtime, reverse=True)
    return archivos[0]

def esperar_archivo_descargado(carpeta_descargas, tiempo_maximo=5, poll_interval=0.8):
    tiempo_inicio = time.time()
    if not carpeta_descargas or not os.path.isdir(carpeta_descargas):
        print(f"Carpeta de descargas inválida: {carpeta_descargas}")
        try:
            mostrar_toast("Carpeta de descargas inválida", tipo="error", titulo="Error")
        except Exception:
            pass
        return None

    print(f"⏳ Esperando archivo en: {carpeta_descargas} (max {tiempo_maximo}s)")
    ultimo_encontrado = None
    while time.time() - tiempo_inicio < tiempo_maximo:
        candidato = obtener_archivo_mas_reciente(carpeta_descargas)
        if candidato and candidato != ultimo_encontrado:
            # comprobar estabilidad y que no sea temporal (.crdownload / .part)
            ruta_final = candidato
            if ruta_final.endswith('.crdownload') or ruta_final.endswith('.part'):
                posible = ruta_final.rsplit('.', 1)[0]
                if os.path.exists(posible):
                    ruta_final = posible
            if os.path.exists(ruta_final) and _archivo_estable(ruta_final):
                print(f"📥 Archivo detectado: {os.path.basename(ruta_final)}")
                try:
                    mostrar_toast(f"Archivo detectado: {os.path.basename(ruta_final)}", tipo="info", titulo="Descarga")
                except Exception:
                    pass
                return ruta_final
            ultimo_encontrado = candidato
        time.sleep(poll_interval)
    print("❌ No se detectó archivo descargado en el tiempo esperado")
    # fallback: buscar EST31100 por si Windows lo listó con retraso
    respaldo = glob.glob(os.path.join(carpeta_descargas, 'EST31100*'))
    if respaldo:
        candidato = max(respaldo, key=os.path.getmtime)
        if _archivo_estable(candidato):
            print(f"⚠️ Usando archivo de respaldo: {os.path.basename(candidato)}")
            return candidato
    return None

def _archivo_estable(ruta):
    """Devuelve True si el archivo existe, tiene tamaño > 0 y puede abrirse para lectura (no bloqueado)."""
    try:
        if not os.path.exists(ruta):
            return False
        tam1 = os.path.getsize(ruta)
        time.sleep(0.8)
        tam2 = os.path.getsize(ruta)
        if tam1 == tam2 and tam1 > 0:
            # Intentar abrir para confirmar que no está bloqueado
            try:
                with open(ruta, 'rb'):
                    pass
                return True
            except Exception:
                return False
        return False
    except Exception:
        return False


def renombrar_y_mover_archivo(archivo_descargado, carpeta_destino, codigo_clinica, indice):
    try:
        if not archivo_descargado or not os.path.exists(archivo_descargado):
            print(f"❌ Archivo descargado no existe: {archivo_descargado}")
            return None

        fecha = datetime.now().strftime("%Y-%m-%d")

        # Nombre bonito según tabla
        nombre_clinica = NOMBRES_CLINICAS.get(codigo_clinica, codigo_clinica)

        nombre_nuevo = f"Informe de la clínica {nombre_clinica} del día {fecha}.xlsx"

        ruta_nueva = os.path.join(carpeta_destino, nombre_nuevo)

        if os.path.exists(ruta_nueva):
            try:
                os.remove(ruta_nueva)
            except Exception as e:
                print(f"⚠️ No se pudo reemplazar el archivo existente: {e}")

        shutil.move(archivo_descargado, ruta_nueva)

        print(f"✅ Archivo renombrado y movido: {nombre_nuevo}")

        try:
            mostrar_toast(f"Archivo movido: {nombre_nuevo}", tipo="success", titulo="Éxito")
        except Exception:
            pass

        return nombre_nuevo

    except Exception as e:
        print(f"❌ Error renombrando/moviendo archivo: {e}")
        return None

def debug_descargas():
    """Función para debug: mostrar qué archivos hay en descargas"""
    carpeta_descargas = obtener_carpeta_descargas()
    print(f"\n🔍 DEBUG - Archivos en descargas:")
    archivos_encontrados = False
    for archivo in os.listdir(carpeta_descargas):
        if 'EST31100' in archivo:
            ruta = os.path.join(carpeta_descargas, archivo)
            tamaño = os.path.getsize(ruta)
            mod_time = time.ctime(os.path.getmtime(ruta))
            print(f"   📄 {archivo} - {tamaño} bytes - {mod_time}")
            archivos_encontrados = True
    
    if not archivos_encontrados:
        print("   ℹ️ No se encontraron archivos EST31100")
    
    # Mostrar todos los archivos Excel también
    print(f"\n🔍 DEBUG - Todos los archivos Excel en descargas:")
    excel_encontrados = False
    for archivo in os.listdir(carpeta_descargas):
        if archivo.endswith(('.xlsx', '.xls', '.síax', '.síac')):
            ruta = os.path.join(carpeta_descargas, archivo)
            tamaño = os.path.getsize(ruta)
            mod_time = time.ctime(os.path.getmtime(ruta))
            print(f"   📊 {archivo} - {tamaño} bytes - {mod_time}")
            excel_encontrados = True
    
    if not excel_encontrados:
        print("   ℹ️ No se encontraron archivos Excel")
# ============================================================
# SISTEMA DE SELECCIÓN DE FECHA PARA INFORMES DE INVENTARIO
# ============================================================

def pedir_fecha_informes():
    global fecha_personalizada

    # ===== CREAR TOPLEVEL =====
    win = ctk.CTkToplevel(root)
    win.title("Seleccionar fecha para los informes")
    win.geometry("480x420")
    win.resizable(False, False)
    win.configure(fg_color="#0F1724")
    win.attributes("-topmost", True)
    win.grab_set()

    # ===== CENTRAR =====
    win.update_idletasks()
    W, H = 480, 420
    x = (win.winfo_screenwidth() // 2) - (W // 2)
    y = (win.winfo_screenheight() // 2) - (H // 2)
    win.geometry(f"{W}x{H}+{x}+{y}")

    # ===== FADE-IN =====
    def fade_in(alpha=0.0):
        if alpha <= 1.0:
            win.attributes("-alpha", alpha)
            win.after(8, lambda: fade_in(alpha + 0.05))
    fade_in()

    # ===== TITULO =====
    ctk.CTkLabel(
        win,
        text="¿Qué fecha deseas usar para los informes?",
        font=("Segoe UI", 20, "bold"),
        text_color="white"
    ).pack(pady=(22, 6))

    ctk.CTkLabel(
        win,
        text="Puedes elegir la fecha de hoy o seleccionar una personalizada.",
        font=("Segoe UI", 14),
        text_color="#AAB4C0"
    ).pack(pady=(0, 15))

    # ===== BOTONES PRINCIPALES =====
    frame_botones = ctk.CTkFrame(win, fg_color="transparent")
    frame_botones.pack(pady=10)

    # Frame donde aparecerá la fecha personalizada
    frame_fecha = ctk.CTkFrame(win, fg_color="transparent")
    frame_fecha.pack(pady=10)

    # ===== FUNCIÓN: USAR HOY =====
    def usar_hoy():
        globals()["fecha_personalizada"] = None
        fade_out_and_close(win)
        pedir_descarga()

    # ===== FUNCIÓN: USAR PERSONALIZADA =====
    def usar_personalizada():

        # Borra contenido previo del área de fecha
        for w in frame_fecha.winfo_children():
            w.destroy()

        # Texto
        ctk.CTkLabel(
            frame_fecha,
            text="Selecciona una fecha:",
            text_color="white",
            font=("Segoe UI", 16, "bold")
        ).pack(pady=6)

        # Selector de fecha
        fecha_widget = DateEntry(
            frame_fecha,
            width=18,
            background="#2C3E50",
            foreground="white",
            borderwidth=2,
            date_pattern="dd-mm-yyyy"
        )
        fecha_widget.pack(pady=6)

        # Guardar fecha
        def guardar_fecha():
            globals()["fecha_personalizada"] = fecha_widget.get()
            fade_out_and_close(win)
            pedir_descarga()

        # Botón confirmar
        ctk.CTkButton(
            frame_fecha,
            text="✔ Continuar con esta fecha",
            fg_color="#3498DB",
            hover_color="#5DADE2",
            corner_radius=14,
            font=("Segoe UI", 16, "bold"),
            height=45,
            width=200,
            command=guardar_fecha
        ).pack(pady=12)

    # ===== BOTÓN HOY =====
    btn_hoy = ctk.CTkButton(
        frame_botones,
        text="Usar fecha HOY",
        fg_color="#16A34A",
        hover_color="#22C55E",
        corner_radius=14,
        font=("Segoe UI", 16, "bold"),
        height=50,
        width=180,
        command=usar_hoy
    )
    btn_hoy.pack(side="left", padx=14)

    # ===== BOTÓN PERSONALIZADA =====
    btn_personalizada = ctk.CTkButton(
        frame_botones,
        text="Fecha PERSONALIZADA",
        fg_color="#2563EB",
        hover_color="#3B82F6",
        corner_radius=14,
        font=("Segoe UI", 16, "bold"),
        height=50,
        width=220,
        command=usar_personalizada
    )
    btn_personalizada.pack(side="left", padx=14)

    # ===== FADE OUT =====
    def fade_out_and_close(toplevel, alpha=1.0):
        if alpha > 0:
            toplevel.attributes("-alpha", alpha)
            toplevel.after(6, lambda: fade_out_and_close(toplevel, alpha - 0.08))
        else:
            toplevel.destroy()

    # ===== PROCESO REAL =====
    def pedir_descarga():
        mostrar_pantalla("inventario")
        root.after(300, descargar_informes_inventario)


def descargar_informes_inventario():
    print("🏁 Iniciando descarga de informes de inventario...")
    mostrar_toast("Iniciando descarga de informes...", tipo="info", titulo="Descarga")
    
    def _proceso_descarga():
        try:
            print("🔗 Abriendo sistema de inventario en Edge...")
            subprocess.Popen(["C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe", URL_RETIRADA])
            time.sleep(12)
            debug_descargas()
            # Crear carpeta específica para esta ejecución
            carpeta_inventario = crear_carpeta_inventario()
            print(f"📁 CARPETA CREADA EN: {carpeta_inventario}")
            
            # Limpiar descargas anteriores
            limpiar_carpeta_descargas()
            
            carpeta_descargas = obtener_carpeta_descargas()
            archivos_descargados = []
            
            # Proceso para descargar 9 informes (uno por cada clínica)
            for i, codigo_clinica in enumerate(CODIGOS_CLINICAS):
                print(f"\n{'='*50}")
                print(f"📥 DESCARGANDO INFORME {i+1}/9 - CLÍNICA {codigo_clinica}")
                print(f"{'='*50}")
                
                # Buscar y hacer clic en el campo de unidad
                if not buscar_y_click("unidad_select.png", "unidad_select", confianza=0.6, intentos=3):
                    print(f"❌ No se pudo encontrar el campo 'unidad' para clínica {codigo_clinica}")
                    continue
                
                try:
                    time.sleep(2)
                    pyautogui.typewrite(codigo_clinica)
                    time.sleep(3)
                    x, y = pyautogui.position()
                    pyautogui.moveTo(x, y + 50)
                    pyautogui.click()
                    print(f"✅ Código de clínica {codigo_clinica} ingresado correctamente.")
                except Exception as e:
                    print(f"❌ Error al ingresar código de clínica {codigo_clinica}: {e}")
                    continue
                
                time.sleep(3)
                
                # Llenar los campos necesarios para el informe
                if not buscar_y_click("tipo_costo.png", "tipo_costo", intentos=3):
                    print(f"❌ No se encontró 'tipo_costo.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("01")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("nivel_totalizacion.png", "nivel_totalizacion", confianza=0.6, intentos=3):
                    print(f"❌ No se encontró 'nivel_totalizacion.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("producto_inicial.png", "producto_inicial", intentos=3):
                    print(f"❌ No se encontró 'producto_inicial.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("1")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("producto_final.png", "producto_final", intentos=3):
                    print(f"❌ No se encontró 'producto_final.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("5")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                if fecha_personalizada:
                    fecha_sin_slash = fecha_personalizada.replace("/", "")
                    buscar_y_click("fecha_campo.png", "campo_fecha", confianza=0.7)
                    pyautogui.press("backspace", presses=10)
                    pyautogui.typewrite(fecha_sin_slash)
                    time.sleep(1)
                # Buscar y hacer clic en el botón de descarga XLSX
                if not buscar_y_click("generar_xlsx.png", "generar_xlsx", confianza=0.9, intentos=3):
                    print(f"❌ No se encontró 'generar_xlsx.png' para clínica {codigo_clinica}")
                    continue
                
                print(f"✅ Descarga iniciada para clínica {codigo_clinica}")
                
                # Esperar a que se complete la descarga
                print("⏳ Esperando a que se complete la descarga...")
                time.sleep(2)
                
                # Buscar el archivo descargado
                archivo_descargado = esperar_archivo_descargado(carpeta_descargas, tiempo_maximo=5)
                
                if archivo_descargado:
                    # Renombrar y mover el archivo
                    archivo_renombrado = renombrar_y_mover_archivo(archivo_descargado, carpeta_inventario, codigo_clinica, i)
                    if archivo_renombrado:
                        archivos_descargados.append(archivo_renombrado)
                        print(f"✅ ÉXITO: Archivo procesado para clínica {codigo_clinica}")
                    else:
                        print(f"❌ FALLO: No se pudo renombrar archivo para clínica {codigo_clinica}")
                else:
                    print(f"❌ FALLO: No se descargó archivo para clínica {codigo_clinica}")
                
                # Recargar la página para la siguiente clínica (solo si no es la última)
                if i < len(CODIGOS_CLINICAS) - 1:
                    print("🔄 Recargando página para siguiente clínica...")
                    try:
                        pyautogui.hotkey("ctrl", "r")
                        time.sleep(7)
                    except Exception:
                        try:
                            pyautogui.press("f5")
                            time.sleep(7)
                        except Exception:
                            pass
            
            # RESUMEN FINAL
            print(f"\n{'='*60}")
            print("🎯 RESUMEN DE DESCARGA")
            print(f"{'='*60}")
            print(f"📁 Carpeta destino: {carpeta_inventario}")
            print(f"📊 Archivos descargados: {len(archivos_descargados)}/9")
            
            if archivos_descargados:
                print("✅ Descarga completada exitosamente")
                mostrar_toast(f"Descarga completada: {len(archivos_descargados)}/9 archivos", tipo="success", titulo="Éxito")
                
                # Guardar la ruta de la carpeta para el procesamiento
                global carpeta_inventario_temporal
                carpeta_inventario_temporal = carpeta_inventario
                
                # Actualizar interfaz con información
                root.after(0, lambda: actualizar_info_inventario(f"Carpeta: {os.path.basename(carpeta_inventario)}", f"Archivos: {len(archivos_descargados)}/9"))
                
                # Habilitar botón de procesar si hay archivos
                if archivos_descargados:
                    root.after(0, lambda: btn_procesar_informes.configure(state="normal"))
            else:
                print("❌ No se descargó ningún archivo")
                mostrar_toast("No se descargaron archivos. Revisa el proceso.", tipo="warning", titulo="Advertencia")
                
        except Exception as e:
            print(f"❌ Error en descarga de informes: {e}")
            mostrar_toast(f"Error en descarga de informes:\n{e}", tipo="error", titulo="Error")
        finally:
            # Reactivar botón
            root.after(0, lambda: btn_descargar_informes.configure(state="normal"))
            print("🔄 Botón 'Descargar Informes' REACTIVADO")
    
    root.after(0, lambda: root.state('zoomed'))
    btn_descargar_informes.configure(state="disabled")
    threading.Thread(target=_proceso_descarga, daemon=True).start()

def procesar_informes_inventario():
    print("🔄 Procesando informes descargados...")
    mostrar_toast("Procesando informes...", tipo="info", titulo="Procesamiento")

    def _proceso():
        try:
            # ===========================
            # 1) Carpeta temporal (descargas ya movidas)
            # ===========================
            if 'carpeta_inventario_temporal' in globals() and os.path.exists(carpeta_inventario_temporal):
                carpeta_temporal = carpeta_inventario_temporal
            else:
                escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
                if not os.path.isdir(escritorio):
                    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")

                carpetas = []
                for item in os.listdir(escritorio):
                    ruta = os.path.join(escritorio, item)
                    if os.path.isdir(ruta) and item.startswith("Informes Inventario de stock"):
                        carpetas.append((ruta, os.path.getmtime(ruta)))

                carpetas.sort(key=lambda x: x[1], reverse=True)
                carpeta_temporal = carpetas[0][0] if carpetas else escritorio

            # ===========================
            # 2) Encontrar archivos descargados
            # ===========================
            archivos_excel = []
            for archivo in os.listdir(carpeta_temporal):
                if archivo.startswith("Informe de la clínica") and archivo.endswith(".xlsx"):
                    archivos_excel.append(os.path.join(carpeta_temporal, archivo))

            if not archivos_excel:
                mostrar_toast("No hay informes para procesar", tipo="warning", titulo="Advertencia")
                return

            resumen = []
            fecha_personalizada_global = globals().get("fecha_personalizada", None)

            def _normalizar_fecha_input(fecha_raw):
                """Acepta DDMMYYYY, DD-MM-YYYY, DD/MM/YYYY, YYYY-MM-DD y devuelve DD-MM-YYYY"""
                if not fecha_raw:
                    return None
                f = str(fecha_raw).strip()
                # formato DDMMYYYY -> insertar guiones
                if re.fullmatch(r"\d{8}", f):
                    return f[0:2] + "-" + f[2:4] + "-" + f[4:8]
                # dd/mm/yyyy o dd-mm-yyyy
                m = re.match(r"(\d{2})[\/\-](\d{2})[\/\-](\d{4})", f)
                if m:
                    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
                # yyyy-mm-dd
                m2 = re.match(r"(\d{4})[\/\-](\d{2})[\/\-](\d{2})", f)
                if m2:
                    return f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"
                return f  # devolver como venga si no coincide

            fecha_inventario_real = None
            if fecha_personalizada_global:
                fecha_inventario_real = _normalizar_fecha_input(fecha_personalizada_global)

            # Si no hay fecha personalizada, extraerla del primer archivo
            if not fecha_inventario_real:
                # intentar extraer del primer excel
                primer_archivo = archivos_excel[0]
                try:
                    df_temp = pd.read_excel(primer_archivo, header=None)
                    for fila in df_temp.values:
                        fila_str = " ".join(map(str, fila))
                        if "Fecha del Inventario" in fila_str:
                            match = re.search(r"(\d{2}[\/\-]\d{2}[\/\-]\d{4})", fila_str)
                            if match:
                                # normalizamos a DD-MM-YYYY
                                fecha_inventario_real = _normalizar_fecha_input(match.group(1))
                                break
                except Exception:
                    fecha_inventario_real = None

            if not fecha_inventario_real:
                fecha_inventario_real = datetime.now().strftime("%d-%m-%Y")

            # ===========================
            # 4) Procesar archivos y extraer totales
            # ===========================
            patrones = ["total del stock", "valor total del stock", "total stock", "stock total"]

            for archivo in archivos_excel:
                try:
                    df = pd.read_excel(archivo, header=None)

                    # Extraer nombre de clínica desde el nombre del archivo
                    nombre_archivo = os.path.basename(archivo)
                    m = re.search(r"Informe de la clínica (.+?) del", nombre_archivo)
                    nombre_clinica = m.group(1).strip() if m else "Desconocida"

                    total_stock = 0
                    encontrado = False

                    for i in range(len(df)):
                        for j in range(len(df.columns)):
                            try:
                                celda = str(df.iat[i, j]).lower()
                            except:
                                celda = ""
                            if any(p in celda for p in patrones):
                                try:
                                    valor_bruto = df.iat[i, j + 2]
                                    try:
                                        valor_float = float(str(valor_bruto).replace(",", "."))
                                    except:
                                        valor_float = 0
                                    total_stock = int(valor_float)

                                except Exception:
                                    total_stock = 0

                                encontrado = True
                                break

                    resumen.append([fecha_inventario_real, nombre_clinica, total_stock, ""])

                except Exception as e:
                    print(f"❌ Error procesando {archivo}: {e}")
                    continue
                
            fecha_hoy = datetime.now().strftime("%d-%m-%Y")
            fecha_inv_fmt = fecha_inventario_real  # ya normalizada como DD-MM-YYYY

            escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
            if not os.path.isdir(escritorio):
                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")

            nombre_carpeta_final = f"Stock {fecha_inv_fmt} {fecha_hoy}"
            carpeta_final = os.path.join(escritorio, nombre_carpeta_final)

            # Si la carpeta ya existe, no crear dentro de otra (asegurar ruta directa)
            if not os.path.exists(carpeta_final):
                os.makedirs(carpeta_final, exist_ok=True)

            # ===========================
            # 6) Mover archivos a la carpeta final (evitar sobreescrituras)
            # ===========================
            for archivo in archivos_excel:
                dst = os.path.join(carpeta_final, os.path.basename(archivo))
                # si existe, renombrar con sufijo _1, _2...
                if os.path.exists(dst):
                    base, ext = os.path.splitext(dst)
                    contador = 1
                    nuevo = f"{base}_{contador}{ext}"
                    while os.path.exists(nuevo):
                        contador += 1
                        nuevo = f"{base}_{contador}{ext}"
                    dst = nuevo
                shutil.move(archivo, dst)

            # ===========================
            # 7) Crear DataFrame final y guardarlo en carpeta_final
            # ===========================
            df_final = pd.DataFrame(resumen, columns=["fecha", "clinica", "inventario", "compra"])

            nombre_salida = f"Resumen Inventario {fecha_inv_fmt}.xlsx"
            ruta_salida = os.path.join(carpeta_final, nombre_salida)
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter

            df_final.to_excel(ruta_salida, index=False)
            wb = openpyxl.load_workbook(ruta_salida)
            ws = wb.active

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")

            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            # quitar decimales del inventario (mostrar como entero)
            for row in range(2, ws.max_row + 1):
                ws[f"C{row}"].number_format = '"$"#,##0'


            # autoajuste columnas
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

            ws.auto_filter.ref = f"A1:D{ws.max_row}"
            wb.save(ruta_salida)

            # ===========================
            # 8) Abrir explorador en la carpeta final
            # ===========================
            try:
                os.startfile(carpeta_final)
            except Exception:
                pass

            print(f"📁 Resumen generado: {ruta_salida}")
            mostrar_toast(f"Resumen generado en {ruta_salida}", tipo="success", titulo="Completado")

        except Exception as e:
            print(f"❌ Error general: {e}")
            mostrar_toast("Error procesando inventario", tipo="error", titulo="Error")
        finally:
            root.after(0, lambda: btn_procesar_informes.configure(state="normal"))

    btn_procesar_informes.configure(state="disabled")
    threading.Thread(target=_proceso, daemon=True).start()



def actualizar_info_inventario(nombre_archivo="", total_items=0):
    try:
        lbl_info_inventario.configure(text=f"Archivo creado: {nombre_archivo}")
        lbl_estado_inventario.configure(text=f"Total de items procesados: {total_items}")
    except Exception as e:
        print(f"Error actualizando info inventario: {e}")

def actualizar_tabla_inventario():
    pass

# FUNCIONES AUXILIARES ORIGINALES 

def buscar_y_click(imagen, nombre, confianza=0.9, intentos=3, esperar=5):
    for i in range(intentos):
        print(f"🔎 Buscando botón '{nombre}' (Intento {i+1}/{intentos})...")
        try:
            ubicacion = pyautogui.locateCenterOnScreen(
                os.path.join(CARPETA_BOTONES, imagen), confidence=confianza
            )
            if ubicacion:
                pyautogui.moveTo(ubicacion, duration=0.2)
                pyautogui.click()
                print(f"✅ Botón '{nombre}' encontrado y clickeado.")
                return True
        except pyautogui.ImageNotFoundException:
            pass
        print("⚠️ No encontrado, reintentando...")
        time.sleep(esperar)
    print(f"❌ No se encontró el botón '{nombre}'.")
    return False

def cargar_datos_desde_supabase():
    global df_facturas, productos_por_factura, codigo_clinica, origen_datos

    try:
        facturas_data = supabase.table("facturas").select("*").execute()
        productos_data = supabase.table("catalogo_productos").select("*").execute()

        print("🧩 Ejemplo de datos en 'productos':")
        if productos_data.data:
            print(productos_data.data[0])
        else:
            print("⚠️ No hay productos en Supabase.")

        if not facturas_data.data:
            mostrar_toast("No se encontraron registros en 'facturas'.", tipo="error", titulo="Sin registros")
            return
        if not productos_data.data:
            mostrar_toast("No se encontraron registros en 'productos'.", tipo="error", titulo="Sin registros")
            return

        facturas = []
        productos_por_factura = {}

        codigo_clinica = str(
            facturas_data.data[0].get("codigo_clinica", "0000")
        ).strip()

        for f in facturas_data.data:
            facturas.append({
            "ID_Factura": str(f["id"]),
            "N° Factura": str(f["numero_factura"]),
            "Fecha": str(f["fecha_factura"]),
            "Empresa": f.get("proveedores", {}).get("nombre", ""),
            "NIT": f.get("proveedores", {}).get("nit", "")
        })


            productos_por_factura[numero_factura] = []

        facturas_por_id = {
            str(f["id"]).strip(): str(f["numero_factura"]).strip()
            for f in facturas_data.data
        }

        for p in productos_data.data:
            factura_id = str(p.get("factura_id", "")).strip()
            if not factura_id:
                continue

            numero_factura = facturas_por_id.get(factura_id)
            if not numero_factura:
                continue 

            if numero_factura not in productos_por_factura:
                productos_por_factura[numero_factura] = []

            productos_por_factura[numero_factura].append(
                {
                    "Código Producto": str(p.get("codigo_producto", "")).strip(),
                    "Nombre Producto": str(p.get("nombre", "")).strip(),
                    "Cantidad": str(p.get("cantidad", "")).replace(",", ".").strip(),
                    "Precio": str(p.get("precio", "")).replace(",", ".").strip(),
                }
            )

        df_facturas = pd.DataFrame(facturas)
        origen_datos = "supabase"

        print(f"✅ Facturas cargadas: {len(facturas)}")
        print("✅ Productos agrupados por factura:")
        for k, v in productos_por_factura.items():
            print(f"  - {k}: {len(v)} productos")

        # --- Actualizar interfaz ---
        actualizar_tabla_facturas()
        try:
            btn_iniciar.configure(state="normal")
            # Animación de confirmación rápida
            quick_pulse_animation(btn_iniciar, "#4CAF50")
        except Exception:
            pass
        try:
            lbl_codigo.configure(text=f"🏥 Datos desde Supabase (Clínica {codigo_clinica})")
        except Exception:
            pass

        mostrar_toast("Datos cargados desde Supabase ✅", tipo="success", titulo="Éxito")

    except Exception as e:
        mostrar_toast(f"No se pudo conectar a Supabase:\n{e}", tipo="error", titulo="Error")

def cargar_datos_desde_excel():
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")],
    )
    if not archivo:
        mostrar_toast("No se seleccionó ningún archivo.", tipo="warning", titulo="Aviso")
        return

    global archivo_excel, df_facturas, productos_por_factura, codigo_clinica, origen_datos
    archivo_excel = archivo

    try:
        df = pd.read_excel(archivo, dtype=str).fillna("")
        columnas_esperadas = [
            "Tipo",
            "N° Factura",
            "Fecha",
            "Empresa",
            "NIT",
            "Código Producto",
            "Nombre Producto",
            "Cantidad",
            "Precio",
        ]
        for col in columnas_esperadas:
            if col not in df.columns:
                mostrar_toast(f"Falta la columna '{col}' en el archivo.", tipo="error", titulo="Error de formato")
                return

        facturas = []
        productos_por_factura = {}
        factura_actual = None

        for _, fila in df.iterrows():
            tipo = str(fila["Tipo"]).strip().upper()

            if tipo == "FACTURA":
                factura_actual = str(fila["N° Factura"]).strip()
                if not factura_actual:
                    continue
                facturas.append(
                    {
                        "ID_Factura": factura_actual,
                        "N° Factura": factura_actual,
                        "Fecha": fila["Fecha"].strip(),
                        "Empresa": fila["Empresa"].strip(),
                        "NIT": fila["NIT"].strip(),
                    }
                )
                productos_por_factura[factura_actual] = []
            elif tipo == "PRODUCTO" and factura_actual:
                productos_por_factura[factura_actual].append(
                    {
                        "Código Producto": str(fila["Código Producto"]).strip(),
                        "Nombre Producto": str(fila["Nombre Producto"]).strip(),
                        "Cantidad": str(fila["Cantidad"]).replace(",", ".").strip(),
                        "Precio": str(fila["Precio"]).replace(",", ".").strip(),
                    }
                )

        df_facturas = pd.DataFrame(facturas)
        codigo_clinica = "".join(
            [c for c in os.path.basename(archivo) if c.isdigit()][-4:]
        )
        if not codigo_clinica:
            codigo_clinica = "0000"

        origen_datos = "excel"
        actualizar_tabla_facturas()
        try:
            btn_iniciar.configure(state="normal")
            quick_pulse_animation(btn_iniciar, "#4CAF50")
        except Exception:
            pass
        try:
            lbl_codigo.configure(text=f"🏥 Código clínica detectado: {codigo_clinica}")
        except Exception:
            pass
        mostrar_toast("Archivo cargado correctamente ✅", tipo="success", titulo="Éxito")

    except Exception as e:
        mostrar_toast(f"No se pudo leer el archivo:\n{e}", tipo="error", titulo="Error")

def actualizar_tabla_facturas():
    for row in tree_facturas.get_children():
        tree_facturas.delete(row)
    for _, fila in df_facturas.iterrows():
        tree_facturas.insert("", "end", values=list(fila))

def mostrar_productos(event):
    item = tree_facturas.selection()
    if not item:
        return
    valores = tree_facturas.item(item[0], "values")
    if len(valores) < 2:
        return
    numero_factura = valores[1]

    productos = productos_por_factura.get(numero_factura, [])
    for row in tree_productos.get_children():
        tree_productos.delete(row)

    for prod in productos:
        tree_productos.insert(
            "",
            "end",
            values=(
                prod["Código Producto"],
                prod["Nombre Producto"],
                prod["Cantidad"],
                prod["Precio"],
            ),
        )

    try:
        lbl_productos.configure(
            text=f"🛒 Productos de la factura seleccionada (Total: {len(productos)})"
        )
    except Exception:
        pass
    print(f"🧾 Factura seleccionada: {numero_factura}")
    print(f"📦 Productos encontrados: {len(productos)}")

def iniciar_proceso():
    if df_facturas.empty:
        mostrar_toast("Primero carga datos desde Excel o Supabase.", tipo="error", titulo="Error")
        return

    print(f"🏁 Iniciando proceso para la clínica {codigo_clinica} ({origen_datos.upper()})...")
    print("🔹 Abriendo TecFood en Edge...")
    subprocess.Popen(["C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe", URL_TECFOOD])
    time.sleep(20)

    # Bucle principal: una iteración por cada fila (factura) en df_facturas
    for index, factura_actual in df_facturas.iterrows():
        numero_factura = str(factura_actual["N° Factura"]).strip()
        empresa = str(factura_actual.get("Empresa", "")).strip()
        nit = str(factura_actual.get("NIT", "")).strip()
        productos = productos_por_factura.get(numero_factura, [])

        print(f"\n\n🔁 Procesando factura {index+1}/{len(df_facturas)} → {numero_factura}")
        print(f"   Empresa: {empresa}  NIT: {nit}  Productos: {len(productos)}")

        if not productos:
            print(f"⚠️ Sin productos para {numero_factura}, se salta.")
            continue
        

        if not buscar_y_click("unidad_select.png", "unidad_select", confianza=0.6):
            mostrar_toast("No se encontró el botón unidad_select.png", tipo="warning", titulo="Botón no encontrado")
            continue

        # Selección de clínica
        try:
            time.sleep(3)
            pyautogui.typewrite(codigo_clinica, interval=0.1)
            time.sleep(2)
            x, y = pyautogui.position()
            pyautogui.moveTo(x, y + 50, duration=0.5)
            pyautogui.click()
            print(f"✅ Clínica {codigo_clinica} seleccionada correctamente.")
        except Exception as e:
            print("Error al seleccionar clínica:", e)
            continue
        pyautogui.press("tab")
        pyautogui.press("tab")
        pyautogui.press("tab")
        time.sleep(2)
        try:
            fecha = pd.to_datetime(factura_actual["Fecha"])
            fecha_formateada = fecha.strftime("%d%m%Y")
            pyautogui.typewrite(fecha_formateada, interval=0.1)
            pyautogui.typewrite(fecha_formateada, interval=0.1)
            pyautogui.press("tab")
            print(f"📅 Fecha ingresada: {fecha_formateada}")
        except Exception as e:
            print("⚠️ Error al procesar la fecha:", e)
            continue

        # Aplicar filtro
        time.sleep(3)
        if not (buscar_y_click("aplicar_filtro.png", "aplicar_filtro") or buscar_y_click("aplicar_filtro_en.png", "aplicar_filtro_en.png")):
            mostrar_toast("No se encontró el botón 'aplicar_filtro'.", tipo="warning", titulo="Botón no encontrado")
            continue

        # Añadir factura
        time.sleep(3)
        if not buscar_y_click("anadir.png", "anadir"):
            mostrar_toast("No se encontró el botón 'anadir.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(4)
        buscar_y_click("seleccionar_archivo.png", "seleccionar")
        time.sleep(1)

        # Buscar y seleccionar PDF correspondiente a esta factura
        try:
            empresa_limpia = "".join(c for c in empresa if c.isalnum() or c.isspace()).lower()
            nit_limpio = "".join(c for c in nit if c.isalnum()).lower()
            carpeta_pdf = r"C:\Users\Maicol Hernandez\Documents\GitHub\PyHealthy\PDF"
            pdf_encontrado: Optional[str] = None
            for archivo in os.listdir(carpeta_pdf):
                nombre_archivo = archivo.lower()
                if (nit_limpio in nombre_archivo and any(palabra in nombre_archivo for palabra in empresa_limpia.split()) and archivo.endswith(".pdf")):
                    pdf_encontrado = os.path.join(carpeta_pdf, archivo)
                    break

            if pdf_encontrado:
                print(f"📄 PDF encontrado: {pdf_encontrado}")
                pyautogui.typewrite(pdf_encontrado)
                time.sleep(1)
                pyautogui.press("enter")
                print("✅ PDF seleccionado correctamente.")
            else:
                mostrar_toast(f"No se encontró un PDF para '{empresa}' y '{nit}'. Verifica carpeta.", tipo="warning", titulo="PDF no encontrado")
                print("⚠️ PDF no encontrado. Se continúa con la factura (si corresponde).")
                # si quieres saltar la factura si no hay PDF, usa: continue
        except Exception as e:
            mostrar_toast(f"No se pudo procesar la factura o buscar el PDF:\n{e}", tipo="error", titulo="Error")
            continue

        # Remitente y número de factura
        time.sleep(5)
        if not buscar_y_click("remitente.png", "remitente"):
            mostrar_toast("No se encontró 'remitente.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        pyautogui.typewrite(nit)
        pyautogui.sleep(3)
        pyautogui.press("tab")
        pyautogui.sleep(7)
        pyautogui.typewrite(numero_factura)
        pyautogui.press("tab")
        pyautogui.sleep(2)

        if not buscar_y_click("serie.png", "serie"):
            mostrar_toast("No se encontró 'serie.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        x, y = pyautogui.position()
        pyautogui.moveTo(x, y + 40, duration=0.5)
        pyautogui.click()
        if not buscar_y_click("emision.png", "emision"):
            mostrar_toast("No se encontró 'emision.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        x, y = pyautogui.position()
        pyautogui.moveTo(x + 50, y + 30, duration=0.5)
        pyautogui.click()
        pyautogui.press("backspace", presses=10)
        pyautogui.typewrite(fecha_formateada)

        if not buscar_y_click("valor.png", "valor"):
            mostrar_toast("No se encontró 'valor.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        pyautogui.typewrite("0")
        pyautogui.press("tab")

        if not buscar_y_click("grabar.png", "grabar"):
            mostrar_toast("No se encontró 'grabar.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        pyautogui_sleep = getattr(pyautogui, "sleep", None)
        if pyautogui_sleep:
            pyautogui_sleep(12)
        else:
            time.sleep(12)

        if not buscar_y_click("productos.png", "productos"):
            mostrar_toast("No se encontró 'productos.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        time.sleep(6)

        # Agregar productos
        for i, producto in enumerate(productos):
            try:
                codigo_producto = str(producto["Código Producto"]).strip()
                cantidad = str(producto["Cantidad"]).strip().replace(",", ".")
                valor_unitario = str(producto["Precio"]).strip().replace(",", ".")

                print(f"➕ Agregando producto {i+1}/{len(productos)}: {codigo_producto}")
                if not buscar_y_click("anadir.png", "añadir"):
                    mostrar_toast("No se encontró 'anadir.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_anadir_no_encontrado")
                time.sleep(11)

                if not buscar_y_click("anadir_producto.png", "añadir_producto"):
                    mostrar_toast("No se encontró 'anadir_producto.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_anadir_producto_no_encontrado")
                time.sleep(11)

                pyautogui.typewrite(codigo_producto)
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40, duration=0.5)
                pyautogui.click()
                time.sleep(3)

                if not buscar_y_click("cantidad.png", "cantidad"):
                    mostrar_toast("No se encontró 'cantidad.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_cantidad_no_encontrado")
                pyautogui.typewrite(cantidad)
                pyautogui.press("tab")
                time.sleep(5)

                pyautogui.typewrite(valor_unitario)
                pyautogui.press("tab")
                time.sleep(5)

                if not buscar_y_click("grabar.png", "grabar"):
                    mostrar_toast("No se encontró 'grabar.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_grabar_no_encontrado")

                print(f"✅ Producto {codigo_producto} grabado correctamente.")
                time.sleep(5)
                pyautogui.press("esc")
                pyautogui_sleep = getattr(pyautogui, "sleep", None)
                if pyautogui_sleep:
                    pyautogui_sleep(5)
                else:
                    time.sleep(5)
            except Exception as e:
                print(f"❌ Error al agregar producto {codigo_producto}: {e}")
                continue

        # Finalizar esta factura
        if not buscar_y_click("FinalizarF.png", "FinalizarF"):
            mostrar_toast("No se encontró 'FinalizarF.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(5)
        if not buscar_y_click("si.png", "si"):
            mostrar_toast("No se encontró 'si.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(2)
        
        print(f"✅ Factura {numero_factura} procesada correctamente (productos: {len(productos)})")

        # --- Recargar la página para la siguiente factura ---
        time.sleep(2)
        try:
            pyautogui.hotkey("ctrl", "r")
            time.sleep(10)
        except Exception:
            try:
                pyautogui.press("f5")
                time.sleep(10)
            except Exception:
                pass
    print("🏁 Proceso completado para todas las facturas.")

# INTERFAZ CON CUSTOMTKINTER
try:
    root
except NameError:
    silenciar_customtkinter()
    root = ctk.CTk()
    root.title("DataSpectra - Sistema de Inventario")
    root.attributes("-fullscreen", True)
    root.configure(fg_color="#0E0F12")

# Sistema simple de multipantallas
contenedor = ctk.CTkFrame(root, fg_color=BG)
contenedor.pack(fill="both", expand=True)

pantallas = {}
def mostrar_pantalla(nombre):
    for p in pantallas.values():
        p.pack_forget()
    pantallas[nombre].pack(fill="both", expand=True)

# PANTALLA PRINCIPAL
pantalla_menu = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["menu"] = pantalla_menu

# Header menu
header_menu = ctk.CTkFrame(pantalla_menu, fg_color=CARD_BG, corner_radius=18)
header_menu.pack(fill="x", padx=20, pady=16)

titulo_menu = ctk.CTkLabel(header_menu, text="DataSpectra — Panel Principal", font=("Segoe UI", 22, "bold"), text_color=TEXT_MAIN)
titulo_menu.pack(side="left", padx=18, pady=10)

# Exit button
def cerrar_app_wrapper():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_exit_menu = ctk.CTkButton(header_menu, text="✕", width=44, height=38, corner_radius=14,
                              fg_color="#FF5252", hover_color="#FF6B6B", font=("Segoe UI", 14, "bold"),
                              command=cerrar_app_wrapper)
btn_exit_menu.pack(side="right", padx=14, pady=8)

# Añadir hover simple
btn_exit_menu.bind("<Enter>", lambda e: simple_button_hover(btn_exit_menu, True))
btn_exit_menu.bind("<Leave>", lambda e: simple_button_hover(btn_exit_menu, False))

content_menu = ctk.CTkScrollableFrame(pantalla_menu, fg_color=BG, corner_radius=0)
content_menu.pack(fill="both", expand=True, padx=20, pady=(10,20))

intro = ctk.CTkLabel(content_menu, text="Selecciona un proceso", font=("Segoe UI", 28, "bold"), text_color=TEXT_MAIN)
intro.pack(pady=(12,6))
sub = ctk.CTkLabel(content_menu, text="Funciones disponibles (las futuras aparecerán habilitadas cuando el administrador las implemente).", font=("Segoe UI", 15), text_color=TEXT_SECOND)
sub.pack(pady=(0,12))

# Lista vertical
def crear_item_lista(parent, icon, titulo, subtitulo, command=None, enabled=True):
    item = ctk.CTkFrame(parent, fg_color="#141518", corner_radius=16)
    item.pack(fill="x", padx=12, pady=8)
    item.configure(height=92)
    item.grid_propagate(False)

    # Icono a la izquierda
    lbl_icon = ctk.CTkLabel(item, text=icon, font=("Segoe UI Emoji", 32), text_color=ACCENT, fg_color="#141518")
    lbl_icon.place(x=18, y=28)

    # Textos
    lbl_tit = ctk.CTkLabel(item, text=titulo, font=("Segoe UI", 15, "bold"), text_color=TEXT_MAIN, fg_color="#141518")
    lbl_tit.place(x=80, y=18)
    lbl_sub = ctk.CTkLabel(item, text=subtitulo, font=("Segoe UI", 11), text_color=TEXT_SECOND, fg_color="#141518")
    lbl_sub.place(x=80, y=44)

    #indicador
    lbl_chev = ctk.CTkLabel(item, text="›", font=("Segoe UI", 22, "bold"), text_color="#8F9AA6", fg_color="#141518")
    lbl_chev.place(relx=0.96, rely=0.5, anchor="e")

    # Hover
    def on_enter(e):
        if enabled:
            simple_item_hover(item, True)
    def on_leave(e):
        if enabled:
            simple_item_hover(item, False)
    item.bind("<Enter>", on_enter)
    item.bind("<Leave>", on_leave)

    # Click
    if enabled and callable(command):
        def onclick(e=None):
            try:
                quick_pulse_animation(item, ACCENT)
                mostrar_toast(f"Abriendo: {titulo}", tipo="info", titulo="Abriendo Operación")
                command()
            except Exception as err:
                mostrar_toast(f"Ocurrió un error: {err}", tipo="error", titulo="Error")
        item.bind("<Button-1>", onclick)
        lbl_icon.bind("<Button-1>", onclick)
        lbl_tit.bind("<Button-1>", onclick)
        lbl_sub.bind("<Button-1>", onclick)
        lbl_chev.bind("<Button-1>", onclick)
    else:
        lbl_tit.configure(text_color="#6F777E")
        lbl_sub.configure(text_color="#4F5559")
        lbl_chev.configure(text_color="#3B4043")

    return item

# Crear 3 items
crear_item_lista(content_menu, "🧾", "Cargar facturas", "Automatiza la carga de facturas en TecFood", command=lambda: mostrar_pantalla("facturas"), enabled=True)
crear_item_lista(content_menu, "📦", "Sistema de Inventario Automatizado", "Descarga y consolida informes de inventario de stock", command=lambda: mostrar_pantalla("inventario"), enabled=True)
crear_item_lista(content_menu, "⚙️", "Función futura #2", "Implementación posterior", command=None, enabled=False)

# footer pequeño
footer_menu = ctk.CTkLabel(content_menu, text="Desarrollado por Area de desarrollo Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
footer_menu.pack(pady=18)

# =========================
# PANTALLA DE FACTURAS
# =========================
pantalla_facturas = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["facturas"] = pantalla_facturas

# === HEADER SUPERIOR ===
header = ctk.CTkFrame(pantalla_facturas, fg_color=CARD_BG, corner_radius=20)
header.pack(fill="x", padx=20, pady=12)

lbl_titulo = ctk.CTkLabel(
    header,
    text="DataSpectra - Carga de Facturas TecFood",
    font=("Segoe UI", 24, "bold"),
    text_color=TEXT_MAIN,
)
lbl_titulo.pack(side="left", padx=20, pady=12)

# Botón volver
def _on_volver():
    mostrar_toast("Volviendo al menú", tipo="info", titulo="Volver")
    mostrar_pantalla("menu")

btn_volver = ctk.CTkButton(
    header,
    text="↩ Volver",
    command=_on_volver,
    width=90,
    height=36,
    corner_radius=16,
    fg_color="#5EA0FF",
    hover_color="#7AB8FF",
    font=("Segoe UI", 13, "bold"),
)
btn_volver.pack(side="right", padx=10, pady=10)

# Añadir hover simple
btn_volver.bind("<Enter>", lambda e: simple_button_hover(btn_volver, True))
btn_volver.bind("<Leave>", lambda e: simple_button_hover(btn_volver, False))

# Botón cerrar
def cerrar_app_facturas():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_close = ctk.CTkButton(
    header,
    text="✕",
    command=cerrar_app_facturas,
    width=40,
    height=40,
    corner_radius=20,
    fg_color="#FF5252",
    hover_color="#FF6B6B",
    font=("Segoe UI", 14, "bold"),
)
btn_close.pack(side="right", padx=10, pady=10)

# Añadir hover simple
btn_close.bind("<Enter>", lambda e: simple_button_hover(btn_close, True))
btn_close.bind("<Leave>", lambda e: simple_button_hover(btn_close, False))

# === CUERPO PRINCIPAL ===
main_frame = ctk.CTkScrollableFrame(pantalla_facturas, fg_color=BG, corner_radius=0)
main_frame.pack(fill="both", expand=True, padx=30, pady=10)

# --- BOTONES DE CARGA ---
btn_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=25)
btn_frame.pack(pady=18)

def _on_cargar_excel():
    mostrar_toast("Selecciona un archivo para cargar", tipo="info", titulo="Cargar Excel")
    quick_pulse_animation(btn_excel, "#FFD54F")
    try:
        cargar_datos_desde_excel()
    except Exception as e:
        mostrar_toast(f"Error al ejecutar cargar_datos_desde_excel: {e}", tipo="error", titulo="Error")

def _on_cargar_supabase():
    mostrar_toast("Cargando desde Supabase...", tipo="info", titulo="Supabase")
    quick_pulse_animation(btn_supabase, "#4FC3F7")
    try:
        cargar_datos_desde_supabase()
    except Exception as e:
        mostrar_toast(f"Error al ejecutar cargar_datos_desde_supabase: {e}", tipo="error", titulo="Error")

btn_excel = ctk.CTkButton(
    btn_frame,
    text="📂 Cargar Excel",
    command=_on_cargar_excel,
    corner_radius=25,
    fg_color=ACCENT,
    hover_color=ACCENT_HOVER,
    text_color="#000000",
    font=("Segoe UI", 14, "bold"),
    width=200,
)
btn_excel.pack(side="left", padx=12, pady=12)

btn_supabase = ctk.CTkButton(
    btn_frame,
    text="🗄️ Cargar desde Supabase",
    command=_on_cargar_supabase,
    corner_radius=25,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 14, "bold"),
    width=250,
)
btn_supabase.pack(side="left", padx=12, pady=12)

btn_excel.bind("<Enter>", lambda e: simple_button_hover(btn_excel, True))
btn_excel.bind("<Leave>", lambda e: simple_button_hover(btn_excel, False))
btn_supabase.bind("<Enter>", lambda e: simple_button_hover(btn_supabase, True))
btn_supabase.bind("<Leave>", lambda e: simple_button_hover(btn_supabase, False))

# --- CÓDIGO DE CLÍNICA ---
lbl_codigo = ctk.CTkLabel(
    main_frame,
    text="Código de clínica detectado: ----",
    font=("Segoe UI", 14, "bold"),
    text_color=TEXT_SECOND,
)
lbl_codigo.pack(pady=6)

# === SECCIÓN DE TABLAS ===
tables_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=20)
tables_frame.pack(fill="both", expand=True, padx=10, pady=10)

def crear_titulo(text):
    return ctk.CTkLabel(
        tables_frame,
        text=text,
        font=("Segoe UI", 16, "bold"),
        text_color=ACCENT,
        anchor="w",
    )

# --- FACTURAS ---
lbl_facturas = crear_titulo("📑 Facturas detectadas")
lbl_facturas.pack(anchor="w", pady=(10, 0), padx=12)

tree_facturas = ttk.Treeview(
    tables_frame,
    columns=("ID_Factura", "N° Factura", "Fecha", "Empresa", "NIT"),
    show="headings",
    height=8,
)
for col in ("ID_Factura", "N° Factura", "Fecha", "Empresa", "NIT"):
    tree_facturas.heading(col, text=col)
    tree_facturas.column(col, anchor="center", width=170)
tree_facturas.pack(fill="x", padx=12, pady=8)
if callable(globals().get("mostrar_productos")):
    tree_facturas.bind("<Double-1>", mostrar_productos)

# --- PRODUCTOS ---
lbl_productos = crear_titulo("🛒 Productos")
lbl_productos.pack(anchor="w", pady=(12, 0), padx=12)

tree_productos = ttk.Treeview(
    tables_frame,
    columns=("Código Producto", "Nombre Producto", "Cantidad", "Precio"),
    show="headings",
    height=8,
)
for col in ("Código Producto", "Nombre Producto", "Cantidad", "Precio"):
    tree_productos.heading(col, text=col)
    tree_productos.column(col, anchor="center", width=170)
tree_productos.pack(fill="x", padx=12, pady=8)

# --- ESTILO DE TABLAS ---
style = ttk.Style()
style.theme_use("clam")
style.configure(
    "Treeview",
    background="#22252A",
    fieldbackground="#22252A",
    foreground="#FFFFFF",
    rowheight=28,
    borderwidth=0,
    font=("Segoe UI", 11),
)
style.configure(
    "Treeview.Heading",
    background="#2E3238",
    foreground="#FFA726",
    font=("Segoe UI", 12, "bold"),
    borderwidth=0,
)
style.map("Treeview", background=[("selected", PRIMARY_HOVER)])

# --- BOTÓN INICIAR ---

def _run_iniciar_proceso_thread():
    try:
        root.after(0, lambda: btn_iniciar.configure(state="disabled"))
        root.after(0, lambda: mostrar_toast("Proceso iniciado en background", tipo="info", titulo="Proceso"))
        iniciar_proceso()
        root.after(0, lambda: btn_iniciar.configure(state="normal"))
        root.after(0, lambda: mostrar_toast("Proceso completado ✅", tipo="success", titulo="Completado"))
        # Animación de éxito rápida
        root.after(0, lambda: quick_pulse_animation(btn_iniciar, "#4CAF50"))
    except Exception as e:
        root.after(0, lambda: mostrar_toast(f"Error en proceso: {e}", tipo="error", titulo="Error"))
        root.after(0, lambda: btn_iniciar.configure(state="normal"))
        # Animación de error rápida
        root.after(0, lambda: quick_pulse_animation(btn_iniciar, "#F44336"))

def _on_iniciar_proceso():
    mostrar_toast("Iniciando proceso...", tipo="info", titulo="Iniciar")
    quick_pulse_animation(btn_iniciar, "#FFD54F")
    t = threading.Thread(target=_run_iniciar_proceso_thread, daemon=True)
    t.start()

btn_iniciar = ctk.CTkButton(
    main_frame,
    text="🚀 Iniciar proceso",
    command=_on_iniciar_proceso,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 15, "bold"),
    corner_radius=24,
    width=250,
    height=48,
    state="disabled",
)
btn_iniciar.pack(pady=16)

btn_iniciar.bind("<Enter>", lambda e: simple_button_hover(btn_iniciar, True))
btn_iniciar.bind("<Leave>", lambda e: simple_button_hover(btn_iniciar, False))

# --- FOOTER ---
footer = ctk.CTkFrame(pantalla_facturas, fg_color=CARD_BG, corner_radius=0)
footer.pack(fill="x", side="bottom")
lbl_footer = ctk.CTkLabel(footer, text="Desarrollado por Area de desarrollo de Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
lbl_footer.pack(pady=8)

# =========================
# PANTALLA DE INVENTARIO
# =========================
pantalla_inventario = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["inventario"] = pantalla_inventario

# === HEADER SUPERIOR ===
header_inventario = ctk.CTkFrame(pantalla_inventario, fg_color=CARD_BG, corner_radius=20)
header_inventario.pack(fill="x", padx=20, pady=12)

lbl_titulo_inventario = ctk.CTkLabel(
    header_inventario,
    text="Sistema de Inventario Automatizado",
    font=("Segoe UI", 24, "bold"),
    text_color=TEXT_MAIN,
)
lbl_titulo_inventario.pack(side="left", padx=20, pady=12)

# Botón volver
def _on_volver_inventario():
    mostrar_toast("Volviendo al menú", tipo="info", titulo="Volver")
    mostrar_pantalla("menu")

btn_volver_inventario = ctk.CTkButton(
    header_inventario,
    text="↩ Volver",
    command=_on_volver_inventario,
    width=90,
    height=36,
    corner_radius=16,
    fg_color="#5EA0FF",
    hover_color="#7AB8FF",
    font=("Segoe UI", 13, "bold"),
)
btn_volver_inventario.pack(side="right", padx=10, pady=10)

btn_volver_inventario.bind("<Enter>", lambda e: simple_button_hover(btn_volver_inventario, True))
btn_volver_inventario.bind("<Leave>", lambda e: simple_button_hover(btn_volver_inventario, False))

# Botón cerrar
def cerrar_app_inventario():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_close_inventario = ctk.CTkButton(
    header_inventario,
    text="✕",
    command=cerrar_app_inventario,
    width=40,
    height=40,
    corner_radius=20,
    fg_color="#FF5252",
    hover_color="#FF6B6B",
    font=("Segoe UI", 14, "bold"),
)
btn_close_inventario.pack(side="right", padx=10, pady=10)

btn_close_inventario.bind("<Enter>", lambda e: simple_button_hover(btn_close_inventario, True))
btn_close_inventario.bind("<Leave>", lambda e: simple_button_hover(btn_close_inventario, False))

# === CUERPO PRINCIPAL ===
main_frame_inventario = ctk.CTkScrollableFrame(pantalla_inventario, fg_color=BG, corner_radius=0)
main_frame_inventario.pack(fill="both", expand=True, padx=30, pady=10)

# --- BOTONES DE PROCESO ---
btn_frame_inventario = ctk.CTkFrame(main_frame_inventario, fg_color=CARD_BG, corner_radius=25)
btn_frame_inventario.pack(pady=18)

def _on_descargar_informes():
    mostrar_toast("Iniciando descarga de informes...", tipo="info", titulo="Descarga")
    quick_pulse_animation(btn_descargar_informes, "#FFD54F")
    try:
        descargar_informes_inventario()
    except Exception as e:
        mostrar_toast(f"Error al descargar informes: {e}", tipo="error", titulo="Error")

btn_descargar_informes = ctk.CTkButton(
    btn_frame_inventario,
    text="📥 Descargar Informes",
    command=pedir_fecha_informes,
    corner_radius=25,
    fg_color=ACCENT,
    hover_color=ACCENT_HOVER,
    text_color="#000000",
    font=("Segoe UI", 14, "bold"),
    width=220,
)
btn_descargar_informes.pack(side="left", padx=12, pady=12)

def _on_procesar_informes():
    mostrar_toast("Procesando informes descargados...", tipo="info", titulo="Procesamiento")
    quick_pulse_animation(btn_procesar_informes, "#4FC3F7")
    try:
        procesar_informes_inventario()
    except Exception as e:
        mostrar_toast(f"Error al procesar informes: {e}", tipo="error", titulo="Error")

btn_procesar_informes = ctk.CTkButton(
    btn_frame_inventario,
    text="🔄 Procesar Informes",
    command=_on_procesar_informes,
    corner_radius=25,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 14, "bold"),
    width=220,
    state="disabled",
)
btn_procesar_informes.pack(side="left", padx=12, pady=12)
def seleccionar_fecha_descarga(root):
    """Ventana popup para elegir si usar fecha de hoy o personalizada"""
    global fecha_personalizada

    ventana = Toplevel(root)
    ventana.title("Seleccionar fecha")
    ventana.geometry("320x180")
    ventana.resizable(False, False)

    Label(ventana, text="¿Qué fecha desea usar para los informes?",
          font=("Arial", 11)).pack(pady=10)

    def usar_hoy():
        nonlocal ventana
        fecha_personalizada = None
        ventana.destroy()

    def usar_personalizada():
        ventana.destroy()
        pedir_fecha_personalizada()

    Button(ventana, text="Usar fecha de HOY", width=25,
           command=usar_hoy, bg="#4CAF50", fg="white").pack(pady=5)

    Button(ventana, text="Elegir FECHA personalizada", width=25,
           command=usar_personalizada, bg="#2196F3", fg="white").pack(pady=5)

    ventana.grab_set()
    root.wait_window(ventana)


def pedir_fecha_personalizada():
    """Ventana para pedir una fecha exacta al usuario"""
    global fecha_personalizada

    top = Toplevel()
    top.title("Elegir Fecha")
    top.geometry("260x150")
    top.resizable(False, False)

    Label(top, text="Seleccione la fecha del informe:",
          font=("Arial", 10)).pack(pady=10)

    fecha_var = StringVar()

    calendario = DateEntry(top, width=18, date_pattern="dd/mm/yyyy",
                           textvariable=fecha_var)
    calendario.pack(pady=5)

    def guardar_fecha():
        nonlocal top
        global fecha_personalizada
        fecha_personalizada = fecha_var.get()  # formato dd/mm/yyyy
        print(f"📌 Fecha personalizada seleccionada: {fecha_personalizada}")
        mostrar_toast(f"Fecha seleccionada: {fecha_personalizada}",
                      tipo="info", titulo="Fecha personalizada")
        top.destroy()

    Button(top, text="Guardar fecha",
           bg="#4CAF50", fg="white", width=20,
           command=guardar_fecha).pack(pady=10)
# Añadir hover simple
btn_descargar_informes.bind("<Enter>", lambda e: simple_button_hover(btn_descargar_informes, True))
btn_descargar_informes.bind("<Leave>", lambda e: simple_button_hover(btn_descargar_informes, False))
btn_procesar_informes.bind("<Enter>", lambda e: simple_button_hover(btn_procesar_informes, True))
btn_procesar_informes.bind("<Leave>", lambda e: simple_button_hover(btn_procesar_informes, False))

# --- INFORMACIÓN DE ARCHIVO Y ESTADO ---
info_frame_inventario = ctk.CTkFrame(main_frame_inventario, fg_color=BG)
info_frame_inventario.pack(pady=12, fill="x")

lbl_info_inventario = ctk.CTkLabel(
    info_frame_inventario,
    text="Esperando descarga de informes...",
    font=("Segoe UI", 14, "bold"),
    text_color=TEXT_SECOND,
)
lbl_info_inventario.pack(anchor="w", pady=(0, 4))

lbl_estado_inventario = ctk.CTkLabel(
    info_frame_inventario,
    text="Estado: Listo para comenzar",
    font=("Segoe UI", 12),
    text_color=TEXT_SECOND,
)
lbl_estado_inventario.pack(anchor="w")

# === SECCIÓN INFORMATIVA ===
info_frame = ctk.CTkFrame(main_frame_inventario, fg_color=CARD_BG, corner_radius=20)
info_frame.pack(fill="x", padx=10, pady=10)

lbl_info_titulo = ctk.CTkLabel(
    info_frame,
    text="📋 Proceso de Inventario Automatizado",
    font=("Segoe UI", 16, "bold"),
    text_color=ACCENT,
    anchor="w",
)
lbl_info_titulo.pack(anchor="w", pady=(10, 5), padx=12)

lbl_info_desc = ctk.CTkLabel(
    info_frame,
    text="Este proceso automatizado:\n• Descarga 9 informes de inventario (uno por cada clínica)\n• Escribe automáticamente los códigos de cada clínica\n• Llena los campos necesarios y descarga en formato XLSX\n• Combina todos los datos en un solo archivo Excel\n• Guarda el resultado como 'Inventario_Combinado_{Fecha_Hora}'",
    font=("Segoe UI", 12),
    text_color=TEXT_SECOND,
    anchor="w",
    justify="left",
)
lbl_info_desc.pack(anchor="w", pady=(0, 10), padx=12)

# --- FOOTER ---
footer_inventario = ctk.CTkFrame(pantalla_inventario, fg_color=CARD_BG, corner_radius=0)
footer_inventario.pack(fill="x", side="bottom")
lbl_footer_inventario = ctk.CTkLabel(footer_inventario, text="Desarrollado por Area de desarrollo de Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
lbl_footer_inventario.pack(pady=8)

# MOSTRAR PANTALLA INICIAL
mostrar_pantalla("menu")

# Solo iniciamos mainloop si no hay otro en el archivo
if not hasattr(root, "_DataSpectra_mainloop_started"):
    root._DataSpectra_mainloop_started = True
    root.mainloop()